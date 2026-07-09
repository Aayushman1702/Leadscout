import re
import tempfile
from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


EXPORT_COLUMNS = [
    ("Name", "name"),
    ("Category", "category"),
    ("Address", "address"),
    ("City", "city"),
    ("State", "state"),
    ("Country", "country"),
    ("Phone", "phone"),
    ("Website", "website"),
    ("Email", "email"),
    ("Instagram", "instagram"),
    ("Facebook", "facebook"),
    ("LinkedIn", "linkedin"),
    ("WhatsApp", "whatsapp"),
    ("Rating", "rating"),
    ("Popularity", "popularity_score"),
    ("Lead Score", "lead_score"),
]


def _field_value(business, field_name):
    value = getattr(business, field_name, "")
    return "" if value is None else value


def _filename_part(value, fallback):
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", str(value or "").strip()).strip("_")
    return cleaned or fallback


def _export_filename(businesses):
    first_business = businesses[0]
    city = _filename_part(_field_value(first_business, "city"), "Leads")
    category = _filename_part(_field_value(first_business, "category"), "Businesses")
    return f"{city}_{category}_{date.today().isoformat()}.xlsx"


def create_businesses_excel(businesses):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Selected Businesses"
    sheet.freeze_panes = "A2"

    headers = [label for label, _ in EXPORT_COLUMNS]
    sheet.append(headers)
    for cell in sheet[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font

    for business in businesses:
        sheet.append([_field_value(business, field_name) for _, field_name in EXPORT_COLUMNS])

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)

    filename = _export_filename(businesses)
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp_file.close()
    workbook.save(temp_file.name)

    return Path(temp_file.name), filename
